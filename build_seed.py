import openpyxl
import json
import re

# ============================================================
# 1. Parse PRR_AllUnits_Properties — has jurisdiction, name, affordability, units
# ============================================================
wb1 = openpyxl.load_workbook('PRR_AllUnits_Properties_20260317.xlsx', read_only=True)
ws1 = wb1['PRR_AllUnits_Properties_2026031']
rows1 = list(ws1.iter_rows(values_only=True))
headers1 = rows1[0]
print("PRR Headers:", headers1)

# Group by (jurisdiction, property_name)
prr_props = {}
for row in rows1[1:]:
    if not row[0]:
        continue
    jurisdiction = str(row[0]).strip()
    name = str(row[1]).strip() if row[1] else ''
    status = str(row[2]).strip() if row[2] else ''
    affordability_type = str(row[3]).strip() if row[3] else ''
    ami_level = str(row[4]).strip() if row[4] else ''
    units = row[5] if row[5] else 0

    key = (jurisdiction, name)
    if key not in prr_props:
        prr_props[key] = {
            'jurisdiction': jurisdiction,
            'name': name,
            'status': status,
            'affordable_units': 0,
            'total_units': 0,
            'ami_levels': [],
        }
    if affordability_type == 'Affordable':
        prr_props[key]['affordable_units'] += int(units) if units else 0
        if ami_level and ami_level not in prr_props[key]['ami_levels']:
            prr_props[key]['ami_levels'].append(ami_level)
    prr_props[key]['total_units'] += int(units) if units else 0

wb1.close()

# Focus on East King County cities
target_cities = ['Bellevue', 'Redmond', 'Kirkland', 'Issaquah', 'Sammamish', 'Mercer Island', 'Kenmore', 'Bothell', 'Woodinville', 'Newcastle', 'Renton']
prr_filtered = {k: v for k, v in prr_props.items() if v['jurisdiction'] in target_cities}
print(f"\nPRR properties in target cities: {len(prr_filtered)}")
for k, v in list(prr_filtered.items())[:5]:
    print(' ', v)

# ============================================================
# 2. Parse Subsidized Rental Housing Inventory — city-level data
# ============================================================
wb2 = openpyxl.load_workbook('Subsidized-Rental-Housing-Inventory-2023.xlsx', read_only=True)
ws_city = wb2['City Inventory']
rows2 = list(ws_city.iter_rows(values_only=True))
print("\nSubsidized City Inventory headers:", rows2[0])
print("Row 1:", rows2[1])
print("Row 2:", rows2[2])

# Find Bellevue row
for r in rows2:
    if r[0] and 'Bellevue' in str(r[0]):
        print("Bellevue row:", r)
wb2.close()

print("\n\nAll PRR properties in Bellevue:")
for k, v in prr_props.items():
    if v['jurisdiction'] == 'Bellevue':
        print(' ', v)

print("\n\nAll PRR properties in Kirkland:")
for k, v in prr_props.items():
    if v['jurisdiction'] == 'Kirkland':
        print(' ', v)

print("\n\nAll PRR properties in Redmond:")
for k, v in prr_props.items():
    if v['jurisdiction'] == 'Redmond':
        print(' ', v)
