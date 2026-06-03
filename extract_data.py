import openpyxl
import json

# ---- 1. PRR_AllUnits_Properties ----
wb1 = openpyxl.load_workbook('PRR_AllUnits_Properties_20260317.xlsx', read_only=True)
ws1 = wb1['PRR_AllUnits_Properties_2026031']
rows1 = list(ws1.iter_rows(values_only=True))
print('PRR Headers:', rows1[0])
print('PRR sample rows:')
for r in rows1[1:6]:
    print(' ', r)
print('PRR total rows:', len(rows1))
wb1.close()

print()

# ---- 2. Subsidized Rental Housing Inventory ----
wb2 = openpyxl.load_workbook('Subsidized-Rental-Housing-Inventory-2023.xlsx', read_only=True)
print('Subsidized sheets:', wb2.sheetnames)

for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f'\n--- Sheet: {sheet_name} ({len(rows)} rows) ---')
    for r in rows[:5]:
        print(' ', r)
wb2.close()
